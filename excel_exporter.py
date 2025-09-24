import pandas as pd
import io
from typing import List, Dict, Any
import streamlit as st
from datetime import datetime
from openpyxl.styles import NamedStyle, Font, PatternFill

class ExcelExporter:
    """Handles exporting candidate data and job data to Excel format"""
    
    def __init__(self):
        pass
    
    def export_candidates(self, candidates_data: List[Dict[str, Any]]) -> bytes:
        """
        Export candidates data to Excel file
        
        Args:
            candidates_data: List of candidate information dictionaries
            
        Returns:
            Excel file as bytes
        """
        try:
            if not candidates_data:
                raise ValueError("No candidate data to export")
            
            # Prepare data for DataFrame
            excel_data = []
            
            for candidate in candidates_data:
                row = {
                    'First Name': candidate.get('first name', ''),
                    'Last Name': candidate.get('last name', ''),
                    'Mobile': candidate.get('mobile', ''),
                    'Email': candidate.get('email', ''),
                    'Current Job Title': candidate.get('current job title', ''),
                    'Current Company': candidate.get('current company', ''),
                    'Previous Job Title': candidate.get('previous job title', ''),
                    'Previous Company': candidate.get('previous company', ''),
                    'Source File': candidate.get('filename', '')
                }
                excel_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Create Excel writer object
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Resume Data', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Resume Data']
                
                # Create a style for headers
                header_style = NamedStyle(name="header_style")
                header_style.font = Font(bold=True)
                #header_style.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                # Register the style to the workbook
                if "header_style" not in workbook.named_styles:
                    workbook.add_named_style(header_style)

                # Apply style to header row
                for cell in worksheet[1]:
                    cell.style = header_style
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set width with some padding, but cap at reasonable maximum
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error creating Excel file: {str(e)}")
            raise e
    
    def export_jobs(self, jobs_data: List[Dict[str, Any]]) -> bytes:
        """
        Export job data to Excel file with Business Nature column
        
        Args:
            jobs_data: List of job information dictionaries
            
        Returns:
            Excel file as bytes
        """
        try:
            if not jobs_data:
                raise ValueError("No job data to export")
            
            # Prepare data for DataFrame
            excel_data = []
            
            for job in jobs_data:
                row = {
                    'Job Title': job.get('Job Title', ''),
                    'Company': job.get('Company', ''),
                    'Business Nature': job.get('Business Nature', ''),
                    'Location': job.get('Location', ''),
                    'Salary': job.get('Salary', ''),
                    'Job URL': job.get('Job URL', '')
                }
                excel_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Create Excel writer object
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Job Data', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Job Data']
                
                # Create a style for headers
                header_style = NamedStyle(name="job_header_style")
                header_style.font = Font(bold=True)
                #header_style.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                # Register the style to the workbook
                if "job_header_style" not in workbook.named_styles:
                    workbook.add_named_style(header_style)

                # Apply style to header row
                for cell in worksheet[1]:
                    cell.style = header_style
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set width with some padding, but cap at reasonable maximum
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error creating Excel file: {str(e)}")
            raise e
